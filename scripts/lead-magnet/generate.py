#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Stokoloji — "Stok Yönetimi Komuta Paneli" lead magnet Excel üreticisi.

Hepsi-bir-arada hesaplama paneli: EOQ + Emniyet Stoğu + Sipariş Noktası (ROP)
+ Stok Devir Hızı. Kullanıcı kendi rakamlarını girer, Excel canlı hesaplar ve
renkli yorum verir. Formüller sitedeki lib/tools/*.ts ile birebir aynı.

Çalıştır:  python3 scripts/lead-magnet/generate.py
Çıktı:     apps/web/public/sablonlar/stokoloji-stok-yonetimi-komuta-paneli.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter

# ── Marka paleti (apps/web/src/app/theme.css) ─────────────────────────────
INK        = "0F2A43"   # lacivert — başlıklar
INK_LIGHT  = "D6E4F0"
TEAL       = "0EA5A4"   # vurgu
TEAL_TINT  = "E7F6F6"   # girdi hücreleri
BG         = "F7FAFC"
BORDER     = "E2E8F0"
WHITE      = "FFFFFF"
GOOD       = "15803D"   # cat-tedarik (yeşil)
GOOD_TINT  = "E6F4EA"
WARN       = "B45309"   # cat-uretim (amber)
WARN_TINT  = "FBF0E1"
DANGER     = "9F1239"   # cat-maliyet (kırmızı)
DANGER_TINT= "F8E5EB"
PURPLE     = "6D28D9"

SITE = "stokoloji.com"

# ── Stil yardımcıları ─────────────────────────────────────────────────────
def font(size=11, bold=False, color=INK, italic=False, name="Calibri"):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)

def fill(color):
    return PatternFill("solid", fgColor=color)

thin = Side(style="thin", color=BORDER)
box = Border(left=thin, right=thin, top=thin, bottom=thin)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right", vertical="center")
WRAP   = Alignment(horizontal="left", vertical="top", wrap_text=True)

CUR = '#,##0 "₺"'      # ₺
NUM = '#,##0.0'
INT = '#,##0'

wb = Workbook()

def title_block(ws, title, subtitle, last_col="F"):
    """Üst marka şeridi + başlık."""
    ws.merge_cells(f"B2:{last_col}2")
    c = ws["B2"]
    c.value = "STOKOLOJI"
    c.font = font(11, bold=True, color=TEAL, name="Consolas")
    c.alignment = LEFT
    ws.merge_cells(f"B3:{last_col}3")
    c = ws["B3"]
    c.value = title
    c.font = font(20, bold=True, color=INK)
    c.alignment = LEFT
    ws.merge_cells(f"B4:{last_col}4")
    c = ws["B4"]
    c.value = subtitle
    c.font = font(11, color="5A6B7B", italic=True)
    c.alignment = LEFT
    ws.row_dimensions[3].height = 28

def footer_link(ws, row, text, last_col="F"):
    ws.merge_cells(f"B{row}:{last_col}{row}")
    c = ws[f"B{row}"]
    c.value = text
    c.font = font(9, color=TEAL)
    c.alignment = LEFT

def section(ws, row, text, col="B", span=2):
    last = get_column_letter(2 + span - 1)
    ws.merge_cells(f"{col}{row}:{last}{row}")
    c = ws[f"{col}{row}"]
    c.value = text
    c.font = font(11, bold=True, color=WHITE)
    c.fill = fill(INK)
    c.alignment = LEFT
    for cc in range(2, 2 + span):
        ws.cell(row=row, column=cc).fill = fill(INK)

def label_cell(ws, row, text, col=2):
    c = ws.cell(row=row, column=col, value=text)
    c.font = font(11, color=INK)
    c.alignment = LEFT
    c.border = box
    c.fill = fill(BG)

def input_cell(ws, row, value, col=4, fmt=NUM):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(12, bold=True, color=INK)
    c.alignment = RIGHT
    c.border = box
    c.fill = fill(TEAL_TINT)
    c.number_format = fmt
    c.protection = Protection(locked=False)   # kullanıcı düzenleyebilir
    return c

def result_cell(ws, row, formula, col=4, fmt=NUM, big=False):
    c = ws.cell(row=row, column=col, value=formula)
    c.font = font(14 if big else 12, bold=True, color=TEAL if big else INK)
    c.alignment = RIGHT
    c.border = box
    c.fill = fill(TEAL_TINT if big else WHITE)
    c.number_format = fmt
    return c

def unit_cell(ws, row, text, col=5):
    c = ws.cell(row=row, column=col, value=text)
    c.font = font(10, color="5A6B7B")
    c.alignment = LEFT
    c.border = box
    c.fill = fill(WHITE)

def base_layout(ws, widths):
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = INK
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    # kenar boşluğu
    ws.column_dimensions["A"].width = 2

# ════════════════════════════════════════════════════════════════════════
# 1) BAŞLA — kapak + nasıl kullanılır
# ════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Başla"
base_layout(ws, {"B": 4, "C": 60, "D": 18, "E": 14, "F": 6})

title_block(ws, "Stok Yönetimi Komuta Paneli",
            "Dört temel stok kararını tek dosyada hesapla ve yorumla.")

intro = ("Bu dosya; sipariş miktarından emniyet stoğuna, yeniden sipariş "
         "noktasından stok devir hızına kadar dört temel stok kararını tek "
         "yerde toplar. Sadece turkuaz hücrelere kendi rakamlarını gir; "
         "panel gerisini hesaplar ve sayıların ne anlama geldiğini Türkçe "
         "yorumlar. Formüller stokoloji.com'daki hesaplayıcılarla birebir aynıdır.")
ws.merge_cells("C6:F9")
ws["C6"].value = intro
ws["C6"].font = font(11, color=INK)
ws["C6"].alignment = WRAP

section(ws, 11, "Nasıl kullanılır", col="C", span=4)
steps = [
    ("1", "Alttaki sekmelerden başla: EOQ, Emniyet Stoğu, Sipariş Noktası, Stok Devir Hızı."),
    ("2", "Yalnızca turkuaz renkli hücrelere kendi verini gir; beyaz hücreler otomatik hesaplanır."),
    ("3", "Her sayfanın altındaki yorum kutusu sonucun ne anlama geldiğini söyler."),
    ("4", "Komuta Paneli sekmesi tüm sonuçları tek bakışta özetler."),
]
r = 12
for n, txt in steps:
    ws.cell(row=r, column=2, value=n).font = font(13, bold=True, color=TEAL)
    ws.cell(row=r, column=2).alignment = CENTER
    ws.merge_cells(f"C{r}:F{r}")
    ws[f"C{r}"].value = txt
    ws[f"C{r}"].font = font(11, color=INK)
    ws[f"C{r}"].alignment = LEFT
    ws.row_dimensions[r].height = 22
    r += 1

section(ws, 18, "Sekmeler", col="C", span=4)
tabs = [
    ("EOQ", "Ekonomik Sipariş Miktarı — siparişi kaç adetlik partiler hâlinde vereceğini bulur."),
    ("Emniyet Stoğu", "Talep ve tedarik dalgalanmasına karşı tutman gereken tampon stok."),
    ("Sipariş Noktası", "Eldeki stok hangi seviyeye düşünce yeni sipariş vermelisin (ROP)."),
    ("Stok Devir Hızı", "Stoğunu yılda kaç kez döndürüyorsun; sermaye verimliliğinin göstergesi."),
]
r = 19
for name, desc in tabs:
    ws[f"C{r}"].value = name
    ws[f"C{r}"].font = font(11, bold=True, color=INK)
    ws[f"C{r}"].alignment = LEFT
    ws.merge_cells(f"D{r}:F{r}")
    ws[f"D{r}"].value = desc
    ws[f"D{r}"].font = font(10, color="5A6B7B")
    ws[f"D{r}"].alignment = LEFT
    ws.row_dimensions[r].height = 30
    r += 1

footer_link(ws, 25, f"Kavramların detaylı anlatımı ve daha fazla hesaplayıcı için: {SITE}")
footer_link(ws, 26, f"© {SITE} — Bu şablonu ekibinle paylaşabilirsin. Hazırlayan: Stokoloji.")

# ════════════════════════════════════════════════════════════════════════
# Hesaplama sayfaları için ortak iskelet
# ════════════════════════════════════════════════════════════════════════
def calc_layout(ws):
    base_layout(ws, {"B": 42, "C": 4, "D": 16, "E": 22, "F": 6})
    ws.column_dimensions["C"].width = 2

# ── 3) EOQ ───────────────────────────────────────────────────────────────
eoq = wb.create_sheet("EOQ")
calc_layout(eoq)
title_block(eoq, "Ekonomik Sipariş Miktarı (EOQ)",
            "Sipariş ve taşıma maliyetini birlikte en aza indiren parti büyüklüğü.")
section(eoq, 6, "Senin verilerin (turkuaz hücreleri doldur)")
label_cell(eoq, 7, "Yıllık talep (D)");            input_cell(eoq, 7, 12000, fmt=INT); unit_cell(eoq, 7, "birim / yıl")
label_cell(eoq, 8, "Sipariş başına maliyet (S)");  input_cell(eoq, 8, 750, fmt=CUR);  unit_cell(eoq, 8, "₺ / sipariş")
label_cell(eoq, 9, "Birim yıllık taşıma maliyeti (H)"); input_cell(eoq, 9, 18, fmt=CUR); unit_cell(eoq, 9, "₺ / birim·yıl")

section(eoq, 11, "Sonuç")
label_cell(eoq, 12, "Ekonomik sipariş miktarı (EOQ)")
result_cell(eoq, 12, "=IFERROR(SQRT((2*D7*D8)/D9),0)", fmt=INT, big=True); unit_cell(eoq, 12, "birim / sipariş")
label_cell(eoq, 13, "Yıllık sipariş sayısı")
result_cell(eoq, 13, "=IFERROR(D7/D12,0)", fmt=NUM); unit_cell(eoq, 13, "sipariş / yıl")
label_cell(eoq, 14, "Siparişler arası süre")
result_cell(eoq, 14, "=IFERROR(365/D13,0)", fmt=INT); unit_cell(eoq, 14, "gün")
label_cell(eoq, 15, "Tahmini yıllık toplam maliyet")
result_cell(eoq, 15, "=IFERROR((D7/D12)*D8+(D12/2)*D9,0)", fmt=CUR); unit_cell(eoq, 15, "₺ / yıl")

section(eoq, 17, "Yorum")
eoq.merge_cells("B18:E20")
eoq["B18"].value = ("Yukarıdaki EOQ, sipariş maliyeti ile taşıma maliyetini birlikte en "
    "aza indiren parti büyüklüğüdür. Daha sık ve küçük sipariş stok maliyetini düşürür "
    "ama sipariş başına maliyeti artırır; EOQ tam denge noktasıdır. 'Yıllık sipariş "
    "sayısı' ve 'siparişler arası süre' satırları, sipariş takvimini planlamana yardımcı olur.")
eoq["B18"].font = font(11, color=INK); eoq["B18"].alignment = WRAP
eoq["B18"].fill = fill(TEAL_TINT)
footer_link(eoq, 22, f"EOQ nedir, ne zaman yetersiz kalır? → {SITE}/blog/eoq-nedir")

# ── 4) Emniyet Stoğu ──────────────────────────────────────────────────────
ss = wb.create_sheet("Emniyet Stoğu")
calc_layout(ss)
title_block(ss, "Emniyet Stoğu",
            "Talep ve tedarik süresi dalgalanmasına karşı tuttuğun tampon stok.")
section(ss, 6, "Senin verilerin (turkuaz hücreleri doldur)")
label_cell(ss, 7, "Ortalama günlük talep");      input_cell(ss, 7, 33, fmt=NUM);  unit_cell(ss, 7, "birim / gün")
label_cell(ss, 8, "En yüksek günlük talep");     input_cell(ss, 8, 55, fmt=NUM);  unit_cell(ss, 8, "birim / gün")
label_cell(ss, 9, "Ortalama tedarik süresi");    input_cell(ss, 9, 7, fmt=NUM);   unit_cell(ss, 9, "gün")
label_cell(ss, 10, "En uzun tedarik süresi");    input_cell(ss, 10, 12, fmt=NUM); unit_cell(ss, 10, "gün")

section(ss, 12, "Sonuç")
label_cell(ss, 13, "Emniyet stoğu")
# Maks–ortalama yöntemi: (maks talep·maks süre) − (ort talep·ort süre)
result_cell(ss, 13, "=MAX(0,(D8*D10)-(D7*D9))", fmt=INT, big=True); unit_cell(ss, 13, "birim")

section(ss, 15, "Yorum")
ss.merge_cells("B16:E18")
ss["B16"].value = ("Emniyet stoğu, talebin veya tedarik süresinin beklenenden kötü gittiği "
    "durumlarda stoksuz kalmamak için tuttuğun tampondur. Yukarıdaki sayı, en kötü senaryo "
    "(en yüksek talep × en uzun tedarik) ile ortalama senaryo arasındaki farkı karşılar. "
    "Talep ve tedarik süresi ne kadar oynaksa bu tampon o kadar büyür.")
ss["B16"].font = font(11, color=INK); ss["B16"].alignment = WRAP
ss["B16"].fill = fill(TEAL_TINT)
ss.merge_cells("B19:E20")
ss["B19"].value = ("Not: Bu basit (maks–ortalama) yöntem istatistik gerektirmez, KOBİ için pratiktir. "
                   "Servis seviyesi (Z-skoru) yöntemini sitedeki hesaplayıcıda bulabilirsin.")
ss["B19"].font = font(9, italic=True, color="5A6B7B"); ss["B19"].alignment = WRAP
footer_link(ss, 22, f"Emniyet stoğu yöntemleri → {SITE}/blog/emniyet-stogu-nedir")

# ── 5) Sipariş Noktası (ROP) ──────────────────────────────────────────────
rop = wb.create_sheet("Sipariş Noktası")
calc_layout(rop)
title_block(rop, "Yeniden Sipariş Noktası (ROP)",
            "Eldeki stok bu seviyeye düşünce yeni sipariş ver.")
section(rop, 6, "Emniyet Stoğu sayfasından otomatik alınır")
label_cell(rop, 7, "Ortalama günlük talep");   result_cell(rop, 7, "='Emniyet Stoğu'!D7", fmt=NUM); unit_cell(rop, 7, "birim / gün")
label_cell(rop, 8, "Ortalama tedarik süresi"); result_cell(rop, 8, "='Emniyet Stoğu'!D9", fmt=NUM); unit_cell(rop, 8, "gün")
label_cell(rop, 9, "Emniyet stoğu");           result_cell(rop, 9, "='Emniyet Stoğu'!D13", fmt=INT); unit_cell(rop, 9, "birim")

section(rop, 11, "Sonuç")
label_cell(rop, 12, "Yeniden sipariş noktası")
result_cell(rop, 12, "=(D7*D8)+D9", fmt=INT, big=True); unit_cell(rop, 12, "birim")
label_cell(rop, 13, "  • Tedarik süresince talep")
result_cell(rop, 13, "=D7*D8", fmt=INT); unit_cell(rop, 13, "birim")
label_cell(rop, 14, "  • Emniyet stoğu payı")
result_cell(rop, 14, "=D9", fmt=INT); unit_cell(rop, 14, "birim")

section(rop, 16, "Yorum")
rop.merge_cells("B17:E19")
rop["B17"].value = ("Eldeki stok, yukarıdaki 'yeniden sipariş noktası' değerine düştüğü anda "
    "yeni sipariş vermelisin. Bu değer iki parçadan oluşur: tedarik süresince tükettiğin "
    "normal miktar (talep × tedarik süresi) + olası gecikme ve ani talep için tuttuğun "
    "emniyet stoğu. Böylece yeni parti gelene kadar stoksuz kalmazsın.")
rop["B17"].font = font(11, color=INK); rop["B17"].alignment = WRAP
rop["B17"].fill = fill(TEAL_TINT)
footer_link(rop, 21, f"ROP nasıl çalışır? → {SITE}/blog/yeniden-siparis-noktasi")

# ── 6) Stok Devir Hızı ────────────────────────────────────────────────────
dev = wb.create_sheet("Stok Devir Hızı")
calc_layout(dev)
title_block(dev, "Stok Devir Hızı",
            "Stoğunu yılda kaç kez döndürdüğün — sermaye verimliliğinin göstergesi.")
section(dev, 6, "Senin verilerin (turkuaz hücreleri doldur)")
label_cell(dev, 7, "Yıllık satılan malın maliyeti (SMM)"); input_cell(dev, 7, 1800000, fmt=CUR); unit_cell(dev, 7, "₺ / yıl")
label_cell(dev, 8, "Dönem başı stok değeri");  input_cell(dev, 8, 380000, fmt=CUR); unit_cell(dev, 8, "₺")
label_cell(dev, 9, "Dönem sonu stok değeri");  input_cell(dev, 9, 420000, fmt=CUR); unit_cell(dev, 9, "₺")

section(dev, 11, "Sonuç")
label_cell(dev, 12, "Ortalama stok değeri")
result_cell(dev, 12, "=(D8+D9)/2", fmt=CUR); unit_cell(dev, 12, "₺")
label_cell(dev, 13, "Stok devir hızı")
result_cell(dev, 13, "=IFERROR(D7/D12,0)", fmt=NUM, big=True); unit_cell(dev, 13, "kez / yıl")
label_cell(dev, 14, "Stok devir gün sayısı")
result_cell(dev, 14, "=IFERROR(365/D13,0)", fmt=INT); unit_cell(dev, 14, "gün")

section(dev, 16, "Yorum")
dev.merge_cells("B17:E21")
dev["B17"].value = ("Stok devir hızı, stoğunu yılda kaç kez döndürdüğünü gösterir; sermaye "
    "verimliliğinin temel ölçüsüdür. Genel referans aralığı:\n"
    "•  4-6 kez: ideal — stok ile satış dengesi sağlıklı.\n"
    "•  4 altı: düşük — sermayen stokta atıl bekliyor, yavaş/fazla ürünleri gözden geçir.\n"
    "•  6-12 kez: yüksek — verimli, ama emniyet stoğunu kontrol et.\n"
    "•  12 üstü: çok yüksek — sık stoksuz kalıyor olabilirsin, sipariş noktanı yükselt.\n"
    "İdeal aralık sektöre göre değişir (gıda/hızlı tüketim yüksek, dayanıklı mal düşük olabilir).")
dev["B17"].font = font(11, color=INK); dev["B17"].alignment = WRAP
dev["B17"].fill = fill(TEAL_TINT)
footer_link(dev, 22, f"İdeal devir hızı sektöre göre nasıl değişir? → {SITE}/blog/stok-devir-hizi")

# ════════════════════════════════════════════════════════════════════════
# 2) KOMUTA PANELİ — tüm sonuçların özeti (Başla'dan sonraya taşınır)
# ════════════════════════════════════════════════════════════════════════
dash = wb.create_sheet("Komuta Paneli")
base_layout(dash, {"B": 4, "C": 34, "D": 20, "E": 4, "F": 34, "G": 20, "H": 4})
title_block(dash, "Komuta Paneli", "Dört stok kararının özeti — tek bakışta.", last_col="G")

def card(ws, r, c_label, label, value_formula, unit, fmt, note_formula):
    """2 sütunluk kart: başlık + büyük değer + not."""
    lc = get_column_letter(c_label)        # başlık sütunu
    vc = get_column_letter(c_label + 1)    # değer sütunu
    ws.merge_cells(f"{lc}{r}:{vc}{r}")
    cc = ws[f"{lc}{r}"]
    cc.value = label
    cc.font = font(11, bold=True, color=WHITE)
    cc.fill = fill(INK); cc.alignment = LEFT
    ws[f"{vc}{r}"].fill = fill(INK)
    # değer
    ws.merge_cells(f"{lc}{r+1}:{vc}{r+1}")
    cv = ws[f"{lc}{r+1}"]
    cv.value = value_formula
    cv.font = font(22, bold=True, color=TEAL)
    cv.alignment = CENTER
    cv.number_format = fmt
    cv.fill = fill(TEAL_TINT)
    ws[f"{vc}{r+1}"].fill = fill(TEAL_TINT)
    ws.row_dimensions[r+1].height = 36
    # birim
    ws.merge_cells(f"{lc}{r+2}:{vc}{r+2}")
    cu = ws[f"{lc}{r+2}"]
    cu.value = unit
    cu.font = font(9, color="5A6B7B"); cu.alignment = CENTER
    cu.fill = fill(TEAL_TINT)
    ws[f"{vc}{r+2}"].fill = fill(TEAL_TINT)
    # not
    ws.merge_cells(f"{lc}{r+3}:{vc}{r+4}")
    cn = ws[f"{lc}{r+3}"]
    cn.value = note_formula
    cn.font = font(10, color=INK); cn.alignment = WRAP
    ws.row_dimensions[r+3].height = 18

# Sol üst — EOQ
card(dash, 6, 3, "Ekonomik Sipariş Miktarı", "=EOQ!D12", "birim / sipariş", INT,
     "En düşük maliyetli parti büyüklüğü")
# Sağ üst — Emniyet Stoğu
card(dash, 6, 6, "Emniyet Stoğu", "='Emniyet Stoğu'!D13", "birim tampon", INT,
     "En kötü senaryo için tampon stok")
# Sol alt — Sipariş Noktası
card(dash, 12, 3, "Yeniden Sipariş Noktası", "='Sipariş Noktası'!D12", "birim", INT,
     "Stok buraya düşünce sipariş ver")
# Sağ alt — Stok Devir Hızı
card(dash, 12, 6, "Stok Devir Hızı", "='Stok Devir Hızı'!D13", "kez / yıl", NUM,
     "İdeal aralık: 4-6 kez / yıl")

dash.merge_cells("C18:G18")
dash["C18"].value = ("Rakamları değiştirmek için ilgili sekmeye git; bu panel otomatik güncellenir. "
                     "Detaylı anlatım ve daha fazla araç: " + SITE)
dash["C18"].font = font(10, italic=True, color="5A6B7B"); dash["C18"].alignment = LEFT

# ── Sayfa sırası: Başla → Komuta Paneli → hesaplayıcılar ──────────────────
wb.move_sheet("Komuta Paneli", -(len(wb.sheetnames) - 2))

OUT_DIR = os.path.join(os.path.dirname(__file__), "..", "..",
                       "apps", "web", "public", "sablonlar")
OUT_DIR = os.path.abspath(OUT_DIR)
os.makedirs(OUT_DIR, exist_ok=True)
OUT = os.path.join(OUT_DIR, "stokoloji-stok-yonetimi-komuta-paneli.xlsx")
wb.save(OUT)

# ── Son-işlem: boş <v></v> önbellek değerlerini temizle ───────────────────
# openpyxl 3.1.5 formül hücrelerine boş <v></v> yazabiliyor; sayı tipli bir
# hücrede boş değer Excel'de "okunamaz içerik" sayılır. Bunları silince formül
# "değeri açılışta hesaplanır" haline gelir ve dosya temiz açılır.
import re as _re, zipfile, shutil, tempfile
_tmp = OUT + ".tmp"
with zipfile.ZipFile(OUT, "r") as zin, \
     zipfile.ZipFile(_tmp, "w", zipfile.ZIP_DEFLATED) as zout:
    for item in zin.infolist():
        data = zin.read(item.filename)
        if item.filename.startswith("xl/worksheets/") and item.filename.endswith(".xml"):
            txt = data.decode("utf-8")
            txt = _re.sub(r"<v\s*/>|<v>\s*</v>", "", txt)
            data = txt.encode("utf-8")
        zout.writestr(item, data)
shutil.move(_tmp, OUT)

print("OK ->", OUT)
print("Sayfalar:", wb.sheetnames)
